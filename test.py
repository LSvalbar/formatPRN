import os
import re
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLineEdit


class FileGroupProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('批量处理PRN文件')
        self.setGeometry(300, 300, 400, 200)

        layout = QVBoxLayout()

        self.file_path_input = QLineEdit(self)
        self.file_path_input.setPlaceholderText("请选择文件夹路径")
        layout.addWidget(self.file_path_input)

        select_folder_button = QPushButton('选择文件夹', self)
        select_folder_button.clicked.connect(self.select_folder)
        layout.addWidget(select_folder_button)

        process_button = QPushButton('开始处理', self)
        process_button.clicked.connect(self.process_files)
        layout.addWidget(process_button)

        close_button = QPushButton('关闭', self)
        close_button.clicked.connect(self.close)
        layout.addWidget(close_button)

        self.setLayout(layout)

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择文件夹")
        if folder:
            self.file_path_input.setText(folder)

    def process_files(self):
        folder_path = self.file_path_input.text().strip()

        if not os.path.exists(folder_path):
            print("文件夹路径无效")
            return

        # 只处理 .prn 文件
        prn_files = [f for f in os.listdir(folder_path) if f.endswith('.prn')]
        if not prn_files:
            print("文件夹中没有找到 PRN 文件")
            return

        # 按文件名分组
        file_groups = self.group_files_by_prefix(prn_files)

        for group_name, files in file_groups.items():
            self.create_excel_for_group(group_name, files, folder_path)

    def group_files_by_prefix(self, file_list):
        file_groups = {}
        for file_name in file_list:
            # 通过正则提取分组前缀
            match = re.match(r"(.+? \d{3})", file_name)
            if match:
                group_name = match.group(1)
                file_groups.setdefault(group_name, []).append(file_name)
        return file_groups

    def create_excel_for_group(self, group_name, file_list, folder_path):
        # 创建新的 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # 定义单元格格式
        cell_style = NamedStyle(name="cell_style", number_format='0.000')

        # 用于存储列标题
        column_titles = {}

        # 用于存储 "相位" 列数据
        phase_data = []

        for file_name in sorted(file_list):
            # 提取列标题 (S11, S12 等)，若包含 "XW"，则为 "相位"
            is_phase = "XW" in file_name
            match = re.search(r"S\d{2}", file_name)
            if not match:
                continue
            column_title = "相位" if is_phase else match.group(0)

            # 如果是 "相位"，存储到专用列表，不立即分配列索引
            if is_phase:
                phase_data.append(file_name)
                continue

            # 分配列索引
            if column_title not in column_titles:
                column_idx = len(column_titles) + 2  # 从第2列开始
                column_titles[column_title] = column_idx
                ws.cell(row=1, column=column_idx, value=column_title)

            # 写入数据
            self.write_file_to_sheet(file_name, ws, column_titles[column_title], folder_path, cell_style)

        # 处理 "相位" 列数据
        if phase_data:
            phase_idx = len(column_titles) + 2  # 放在最后一列
            ws.cell(row=1, column=phase_idx, value="相位")
            for file_name in phase_data:
                self.write_file_to_sheet(file_name, ws, phase_idx, folder_path, cell_style)

        # 保存文件
        output_file = os.path.join(folder_path, f"{group_name}.xlsx")
        wb.save(output_file)
        print(f"Excel 文件已保存: {output_file}")

    def write_file_to_sheet(self, file_name, sheet, column_idx, folder_path, cell_style):
        file_path = os.path.join(folder_path, file_name)
        with open(file_path, 'rb') as f:
            for row_idx, line in enumerate(f, start=2):
                line = line.decode('utf-8', errors='ignore').strip()
                if line:
                    parts = line.split(',')
                    if len(parts) == 2:
                        # 写入A列
                        if not sheet.cell(row=row_idx, column=1).value:
                            sheet.cell(row=row_idx, column=1, value=parts[0].strip())
                        # 写入对应列
                        sheet.cell(row=row_idx, column=column_idx, value=float(parts[1].strip())).style = cell_style


if __name__ == '__main__':
    app = QApplication([])
    window = FileGroupProcessor()
    window.show()
    app.exec_()
